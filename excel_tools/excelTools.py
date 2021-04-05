import time
import openpyxl

class ExcelFilter:
    @staticmethod
    def filterRows(nameOfExcel, createdExcelName, listOfStringToBeFiltered, rowFromFilterStart, masterColumnForFiltering):
        statingTime = time.perf_counter()
        excelFile = openpyxl.load_workbook(nameOfExcel)
        excelSheet = excelFile[excelFile.active.title]
        totalRowsInGivenExcel = excelSheet.max_row

        totalRowsRemovedFromExcel = 0
        setOfRemovedValues = set()

        for i in range(rowFromFilterStart, totalRowsInGivenExcel+1):
            if excelSheet.cell(row=i, column=masterColumnForFiltering).value not in listOfStringToBeFiltered:
                setOfRemovedValues.add(excelSheet.cell(row=i, column=masterColumnForFiltering).value)

        i = rowFromFilterStart
        while i <= excelSheet.max_row:
            if excelSheet.cell(row=i, column=masterColumnForFiltering).value not in listOfStringToBeFiltered:
                excelSheet.delete_rows(i)
                print(f'Deleted row {i}.')
                totalRowsRemovedFromExcel += 1
                
            else:
                i += 1
            
        
        if totalRowsRemovedFromExcel > 0:
            excelFile.save(createdExcelName)
            print(f'Total rows in excel are {totalRowsInGivenExcel} & {totalRowsRemovedFromExcel} rows are removed.')
            print(f"New excel is created in same folder with name '{createdExcelName}'.")
            print(f"List of Unique Values of removed rows present in excel.\n{list(setOfRemovedValues)}")
            
            endingTime = time.perf_counter()
            totalTime = endingTime-statingTime
            print(f'\nTotal time taken is {round(totalTime, 2)} seconds.')
                      
        elif totalRowsRemovedFromExcel == 0:
            print(f'Given excel is already filtered.')
            
        else:
            print('Something went wrong.')
            
            
    @staticmethod
    def getUniqueValueListOfColumn(pathOfExcel, columnToBeSearched):
        excelFile = openpyxl.load_workbook(pathOfExcel)
        excelSheet = excelFile[excelFile.active.title]
        totalRows = excelSheet.max_row

        uniqueCellValueList = []

        for i in range(2, totalRows+1):
            cellValue = excelSheet.cell(row=i, column=columnToBeSearched).value
            if cellValue not in uniqueCellValueList:
                uniqueCellValueList.append(cellValue)
                
        print(f'The Rows of given column has total {len(uniqueCellValueList)} unique values.\n')
        for industry in uniqueCellValueList:
            print(f'{industry}')
        print('')
            
        
            
class WrongPhoneNumberFinder:
    @staticmethod
    def findWrongPhoneNumberRow(nameOfExcel, columnOfPhoneNumber):
        
        excelFile = openpyxl.load_workbook(nameOfExcel)
        excelSheet = excelFile[excelFile.active.title]
        totalRows = excelSheet.max_row

        invalidPhoneNumbersInExcel = 0
        listOfRowsOfInvalidPhoneNumbers = []

        for i in range(2, totalRows+1):
            if len(str(excelSheet.cell(row=i, column=columnOfPhoneNumber).value)) > 10 or len(str(excelSheet.cell(row=i, column=columnOfPhoneNumber).value)) < 10:
                invalidPhoneNumbersInExcel += 1
                listOfRowsOfInvalidPhoneNumbers.append(i)
                wrongNumberValue = excelSheet.cell(row=i, column=columnOfPhoneNumber).value
                print(f"Row {i} has wrong number that is '{wrongNumberValue}'. Kindly check row {i}.")
        
        if invalidPhoneNumbersInExcel == 0:    
            print(f'There is no invalid phone number in the given excel.')
        
        elif invalidPhoneNumbersInExcel > 0:     
            print(f"\nTotal no. of invalid phone number is '{invalidPhoneNumbersInExcel}'. List of Rows having wrong number in excel is given below.\n{listOfRowsOfInvalidPhoneNumbers}")
            
        else:
            print('Something went wrong.')
            

class Report:
    @staticmethod
    def getSummary(nameOfExcel, columnInWhichWeCount): 
        excelFile = openpyxl.load_workbook(nameOfExcel)
        excelSheet = excelFile[excelFile.active.title]

        totalRows = excelSheet.max_row

        totalSMSCount = 0
        deliveredCounts = 0
        failedCounts = 0
        sentCounts = 0
        unknownCounts = 0

        uniqueValuesListPresentInGivenColumn = []

        for i in range(2, totalRows+1):
            valuesInCell = excelSheet.cell(row=i, column=columnInWhichWeCount).value
            if valuesInCell not in uniqueValuesListPresentInGivenColumn:
                uniqueValuesListPresentInGivenColumn.append(valuesInCell)
                
            if valuesInCell == 'Delivered':
                deliveredCounts += 1
            
            elif valuesInCell == 'Failed':
                failedCounts += 1
                
            elif valuesInCell == 'Sent':
                sentCounts += 1
            else:
                unknownCounts += 1
             
        totalSMSCount = deliveredCounts + failedCounts + sentCounts + unknownCounts 

        print(f"Total Unique Values in given column is '{uniqueValuesListPresentInGivenColumn}'.")
        print(f"Total 'SMS' counts is {totalSMSCount}.")
        print(f"Total Messages 'Delivered' is {deliveredCounts}.")
        print(f"Total Messages 'Failed' is {failedCounts}.")
        print(f"Total Messages 'Sent' is {sentCounts}.")
        print(f"Total 'Unknown' counts is {unknownCounts}.")
            